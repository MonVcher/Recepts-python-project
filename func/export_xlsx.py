from openpyxl import load_workbook
# from datetime import datetime as dt
from data import db_session
from data.Recepts import Recept
# db_sess = db_session.create_session()
# path = "export/exp.xlsx"

def extand_xlsx_file(db_sess, file_href) -> str:
    wookbook = load_workbook(file_href)
    worksheet, dataup = wookbook.active, []
    for row in worksheet.iter_rows(2, worksheet.max_row):
        c = []
        for i in range(0, worksheet.max_column):
            if i == 0:
                c.append((str(row[i].value).split())[0])
                print((str(row[i].value).split())[0], end=" ")
            else:
                c.append(str(row[i].value))
                print(row[i].value, end=" ")
        dataup.append(c)
        print('')
    # if db_sess:
    #     students = db_sess.query(Recept)
    # else: 
    #     return False
    for i in dataup:
        visit = Recept(name=i[0],
                              products=i[1],
                              recept=i[2],
                              )
        db_sess.add(visit)
    # return str(dt.today()).split()[0]


# some = extand_xlsx_file(db_sess, path)
# if some:
#     print(some)